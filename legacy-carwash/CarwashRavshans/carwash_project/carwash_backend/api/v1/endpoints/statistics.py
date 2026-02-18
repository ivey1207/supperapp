from fastapi import APIRouter, Depends, HTTPException, Response, Query, Body
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from typing import List, Optional
import datetime
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment

from carwash_backend.db import schemas, repository, models
from carwash_backend.db.database import get_db
from carwash_backend.api.v1.dependencies import get_current_admin

router = APIRouter()

PAYMENT_TYPE_TRANSLATIONS = {
    "cash": "Наличные",
    "online": "Онлайн",
    "rfid_wash": "RFID мойка",
    "rfid_top_up": "RFID пополнение",
    "bonus": "Бонусы",
    "click": "Click",
    "payme": "Payme",
    "uzum": "Uzum"
}

SERVICE_TYPE_TRANSLATIONS = {
    "wash": "Мойка",
    "vacuum": "Пылесос",
    "air_compressor": "Компрессор",
    "water": "Вода"
}

@router.get("/", response_model=schemas.PaginatedTransactionsResponse, summary="Get All Transactions with Pagination")
async def get_all_transactions(
    start_date: Optional[str] = Query(None, description="Дата начала в формате YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="Дата окончания в формате YYYY-MM-DD"),
    device_id: Optional[str] = Query(None, description="ID устройства"),
    payment_type: Optional[str] = Query(None, description="Тип платежа (cash, online, rfid_wash, rfid_top_up)"),
    skip: int = Query(0, ge=0, description="Пропустить записей (для пагинации)"),
    limit: int = Query(10, ge=1, le=100, description="Количество записей на страницу (макс. 100)"),
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):

    try:

        query = select(models.Transaction)

        conditions = []

        if start_date:
            try:
                start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
                conditions.append(models.Transaction.timestamp >= start_dt)
            except ValueError:
                raise HTTPException(status_code=400, detail="Неправильный формат start_date. Используйте YYYY-MM-DD")

        if end_date:
            try:
                end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                conditions.append(models.Transaction.timestamp <= end_dt)
            except ValueError:
                raise HTTPException(status_code=400, detail="Неправильный формат end_date. Используйте YYYY-MM-DD")

        payment_transactions = []
        total_count = 0

        if payment_type:

            valid_types = ["cash", "online", "rfid_wash", "rfid_top_up", "rfid", "bonus", "click", "payme", "uzum"]
            if payment_type not in valid_types:
                raise HTTPException(status_code=400, detail=f"Неверный тип платежа. Доступные: {', '.join(valid_types)}")

            if payment_type in ["click", "payme", "uzum"]:

                payment_conditions = []

                if start_date:
                    try:
                        start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
                        payment_conditions.append(models.PaymentTransaction.created_at >= start_dt)
                    except ValueError:
                        raise HTTPException(status_code=400, detail="Неправильный формат start_date. Используйте YYYY-MM-DD")

                if end_date:
                    try:
                        end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                        payment_conditions.append(models.PaymentTransaction.created_at <= end_dt)
                    except ValueError:
                        raise HTTPException(status_code=400, detail="Неправильный формат end_date. Используйте YYYY-MM-DD")

                payment_conditions.append(models.PaymentTransaction.payment_type == payment_type)

                payment_conditions.append(models.PaymentTransaction.status == "completed")

                payment_query = select(models.PaymentTransaction)
                if payment_conditions:
                    payment_query = payment_query.where(and_(*payment_conditions))
                payment_query = payment_query.order_by(models.PaymentTransaction.created_at.desc())

                payment_count_query = select(func.count()).select_from(payment_query.subquery())
                payment_count_result = await db.execute(payment_count_query)
                total_count = payment_count_result.scalar()

                payment_query = payment_query.offset(skip).limit(limit)

                payment_result = await db.execute(payment_query)
                payment_transactions = payment_result.scalars().all()

                transactions = []
            else:

                if payment_type == "rfid":

                    conditions.append(models.Transaction.type.in_([models.TransactionTypeEnum.rfid_wash, models.TransactionTypeEnum.rfid_top_up]))
                else:
                    conditions.append(models.Transaction.type == payment_type)
        else:

            payment_conditions = []

            if start_date:
                try:
                    start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
                    payment_conditions.append(models.PaymentTransaction.created_at >= start_dt)
                except ValueError:
                    raise HTTPException(status_code=400, detail="Неправильный формат start_date. Используйте YYYY-MM-DD")

            if end_date:
                try:
                    end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                    payment_conditions.append(models.PaymentTransaction.created_at <= end_dt)
                except ValueError:
                    raise HTTPException(status_code=400, detail="Неправильный формат end_date. Используйте YYYY-MM-DD")

            payment_conditions.append(models.PaymentTransaction.status == "completed")

            payment_query = select(models.PaymentTransaction)
            if payment_conditions:
                payment_query = payment_query.where(and_(*payment_conditions))
            payment_query = payment_query.order_by(models.PaymentTransaction.created_at.desc())

            payment_result = await db.execute(payment_query)
            payment_transactions = payment_result.scalars().all()

        if conditions:
            query = query.where(and_(*conditions))

        query = query.order_by(models.Transaction.timestamp.desc())

        if payment_type and payment_type in ["click", "payme", "uzum"]:

            transactions = []

        else:

            if payment_type:

                count_query = select(func.count()).select_from(query.subquery())
                count_result = await db.execute(count_query)
                old_total_count = count_result.scalar()

                query = query.offset(skip).limit(limit)

                result = await db.execute(query)
                transactions = result.scalars().all()
                total_count = old_total_count
            else:

                result = await db.execute(query)
                transactions = result.scalars().all()

        response_data = []

        for transaction in transactions:

            device_name = "system"
            if transaction.type.value == "cash":
                device_name = "bill_acceptor"
            elif transaction.type.value == "online":
                device_name = "payment_terminal"
            elif transaction.type.value in ["rfid_wash", "rfid_top_up"]:
                device_name = "rfid_scanner"

            if device_id and device_name != device_id:
                continue

            response_data.append({
                "id": transaction.id,
                "created_at": transaction.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                "type": PAYMENT_TYPE_TRANSLATIONS.get(transaction.type.value, transaction.type.value),
                "amount": transaction.amount,
                "device_id": device_name,
                "description": transaction.description or f"{PAYMENT_TYPE_TRANSLATIONS.get(transaction.type.value, transaction.type.value)} платеж"
            })

        for payment_transaction in payment_transactions:

            device_name = "payment_terminal"

            if device_id and device_name != device_id:
                continue

            amount_sum = payment_transaction.amount / 100.0 if payment_transaction.amount else 0

            payment_type_display = PAYMENT_TYPE_TRANSLATIONS.get(payment_transaction.payment_type, payment_transaction.payment_type)

            response_data.append({
                "id": payment_transaction.id,
                "created_at": payment_transaction.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "type": payment_type_display,
                "amount": amount_sum,
                "device_id": device_name,
                "description": f"{payment_type_display} платеж - {payment_transaction.transaction_id}"
            })

        if not payment_type:

            response_data.sort(key=lambda x: x['created_at'], reverse=True)

            total_count = len(response_data)
            start_idx = skip
            end_idx = skip + limit
            response_data = response_data[start_idx:end_idx]

        total_pages = (total_count + limit - 1) // limit
        current_page = (skip // limit) + 1

        return {
            "transactions": response_data,
            "pagination": {
                "total_count": total_count,
                "total_pages": total_pages,
                "current_page": current_page,
                "page_size": limit,
                "has_next": skip + limit < total_count,
                "has_previous": skip > 0
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения транзакций: {str(e)}")

@router.get("", response_model=schemas.PaginatedTransactionsResponse, include_in_schema=False)
async def get_all_transactions_alias(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    device_id: Optional[str] = Query(None),
    payment_type: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(10, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):
    return await get_all_transactions(start_date, end_date, device_id, payment_type, skip, limit, db, current_admin)

@router.post("/", response_model=schemas.StatisticsResponse, summary="Get Statistics")
async def get_statistics(
    stats_request: schemas.StatisticsRequest,
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):

    try:
        start_date = datetime.datetime.strptime(stats_request.start_date, "%Y-%m-%d")
        end_date = datetime.datetime.strptime(stats_request.end_date, "%Y-%m-%d")

        end_date = end_date.replace(hour=23, minute=59, second=59)

    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Неправильный формат даты. Используйте YYYY-MM-DD"
        )

    stats = await repository.get_statistics(
        db,
        start_date=start_date,
        end_date=end_date,
        post_id=stats_request.post_id
    )

    return stats

@router.get("/breakdown", summary="Get Payment Breakdown by Provider")
async def get_payment_breakdown(
    start_date: str,
    end_date: str,
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):

    try:
        start = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.datetime.strptime(end_date, "%Y-%m-%d")
        end = end.replace(hour=23, minute=59, second=59)

    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Неправильный формат даты. Используйте YYYY-MM-DD"
        )

    breakdown = await repository.get_payment_summary_by_methods(db, start, end)

    return {
        "period": {
            "start_date": start_date,
            "end_date": end_date
        },
        "total": breakdown["total"],
        "cash": breakdown["cash"],
        "online": breakdown["online"],
        "rfid": breakdown["rfid"],
        "by_provider": breakdown["by_provider"]
    }

@router.post("/export/", summary="Export Data")
async def export_data(
    export_request: schemas.ExportRequest,
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):

    try:
        start_date = datetime.datetime.strptime(export_request.start_date, "%Y-%m-%d")
        end_date = datetime.datetime.strptime(export_request.end_date, "%Y-%m-%d")
        end_date = end_date.replace(hour=23, minute=59, second=59)

    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Неправильный формат даты. Используйте YYYY-MM-DD"
        )

    if export_request.format not in ["excel", "csv"]:
        raise HTTPException(status_code=400, detail="Поддерживаются только форматы: excel, csv")

    if export_request.data_type not in ["transactions", "statistics", "rfid_cards"]:
        raise HTTPException(
            status_code=400,
            detail="Поддерживаются типы данных: transactions, statistics, rfid_cards"
        )

    if export_request.data_type == "transactions":
        data = await repository.get_transactions_for_export(db, start_date, end_date)
        filename = f"transactions_{export_request.start_date}_{export_request.end_date}"

    elif export_request.data_type == "statistics":
        data = await repository.get_statistics_for_export(db, start_date, end_date)
        filename = f"statistics_{export_request.start_date}_{export_request.end_date}"

    elif export_request.data_type == "rfid_cards":
        data = await repository.get_rfid_cards_for_export(db)
        filename = f"rfid_cards_{datetime.datetime.now().strftime('%Y-%m-%d')}"

    if export_request.format == "excel":
        file_content = _create_excel_file(data, export_request.data_type)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename += ".xlsx"
    else:
        file_content = _create_csv_file(data)
        media_type = "text/csv"
        filename += ".csv"

    return Response(
        content=file_content,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/excel", summary="Export Excel Test")
async def export_excel_test(
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):

    try:

        end_date = datetime.datetime.now()
        start_date = end_date - datetime.timedelta(days=7)

        data = await repository.get_transactions_for_export(db, start_date, end_date)
        filename = f"test_export_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"

        file_content = _create_excel_file(data, "transactions")

        return Response(
            content=file_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка экспорта: {str(e)}")

def _create_excel_file(data: List[dict], data_type: str) -> bytes:

    wb = openpyxl.Workbook()
    ws = wb.active

    if not data:
        ws.cell(row=1, column=1, value="Нет данных для экспорта")

    else:

        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row_idx, item in enumerate(data, 2):
            for col_idx, value in enumerate(item.values(), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    if data_type == "transactions":
        ws.title = "Транзакции"
    elif data_type == "statistics":
        ws.title = "Статистика"
    elif data_type == "rfid_cards":
        ws.title = "RFID карты"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def _create_csv_file(data: List[dict]) -> bytes:

    import csv
    from io import StringIO

    if not data:
        return "Нет данных для экспорта".encode('utf-8')

    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)

    return output.getvalue().encode('utf-8')

@router.get("/daily-summary/", summary="Get Daily Summary")
async def get_daily_summary(
    date: str,
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):

    try:
        target_date = datetime.datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Неправильный формат даты. Используйте YYYY-MM-DD"
        )

    summary = await repository.get_daily_summary(db, target_date)
    return summary

@router.get("/daily-summary", include_in_schema=False)
async def get_daily_summary_alias(
    date: str,
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):
    return await get_daily_summary(date, db, current_admin)

@router.get("/posts-activity/", summary="Get Posts Activity")
async def get_posts_activity(
    start_date: Optional[str] = Query(None, description="Дата начала в формате YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="Дата окончания в формате YYYY-MM-DD"),
    hours: Optional[int] = Query(24, description="Количество часов (если не указаны даты)"),
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):

    start_dt, end_dt = None, None
    if start_date and end_date:
        try:
            start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        except ValueError:
            raise HTTPException(status_code=400, detail="Неправильный формат даты. Используйте YYYY-MM-DD")

    final_hours = hours if not (start_dt and end_dt) else None

    if final_hours and (final_hours <= 0 or final_hours > 168):
        raise HTTPException(
            status_code=400,
            detail="Количество часов должно быть от 1 до 168"
        )

    activity = await repository.get_posts_activity(db, start_date=start_dt, end_date=end_dt, hours=final_hours)
    return {
        "start_date": start_date,
        "end_date": end_date,
        "hours": final_hours,
        "posts": activity
    }

@router.get("/posts-activity", include_in_schema=False)
async def get_posts_activity_alias(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    hours: Optional[int] = Query(24),
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):
    return await get_posts_activity(start_date, end_date, hours, db, current_admin)

@router.get("/clients", summary="Get Client Statistics")
async def get_client_statistics(
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):

    import datetime

    now = datetime.datetime.now()

    current_week_start = now - datetime.timedelta(days=now.weekday())
    current_week_start = current_week_start.replace(hour=0, minute=0, second=0, microsecond=0)

    total_clients_query = select(func.count(models.RfidCard.id))
    total_clients_result = await db.execute(total_clients_query)
    total_clients = total_clients_result.scalar() or 0

    active_clients_query = select(func.count(func.distinct(models.WashSession.rfid_card_id))).where(
        and_(
            models.WashSession.started_at >= current_week_start,
            models.WashSession.rfid_card_id.isnot(None)
        )
    )
    active_clients_result = await db.execute(active_clients_query)
    active_clients_this_week = active_clients_result.scalar() or 0

    cards_count_query = select(func.count(models.RfidCard.id))
    cards_count_result = await db.execute(cards_count_query)
    total_cards = cards_count_result.scalar() or 0

    new_clients_this_week = max(1, int(total_cards * 0.1)) if total_cards > 0 else 0

    return {
        "total_clients": total_clients,
        "active_clients_this_week": active_clients_this_week,
        "new_clients_this_week": new_clients_this_week,
        "period_info": {
            "current_week_start": current_week_start.strftime("%Y-%m-%d"),
            "current_week_end": now.strftime("%Y-%m-%d"),
            "last_month_start": (now - datetime.timedelta(days=30)).strftime("%Y-%m-%d"),
            "last_month_end": now.strftime("%Y-%m-%d"),
            "note": "Новые клиенты рассчитаны примерно (10% от всех карт)"
        }
    }

@router.get("/revenue", summary="Get Revenue Statistics")
async def get_revenue_statistics(
    start_date: str,
    end_date: str,
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):

    try:
        start = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.datetime.strptime(end_date, "%Y-%m-%d")
        end = end.replace(hour=23, minute=59, second=59)

        sessions = await repository.get_sessions_by_date_range(db, start, end)

        total_revenue = 0.0
        session_count = len(sessions)

        for session in sessions:
            if session.total_balance:
                total_revenue += session.total_balance

        return {
            "start_date": start_date,
            "end_date": end_date,
            "total_revenue": total_revenue,
            "session_count": session_count,
            "average_session": round(total_revenue / session_count, 2) if session_count > 0 else 0
        }

    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный формат даты. Используйте YYYY-MM-DD")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения статистики: {str(e)}")

@router.get("/summary", summary="Get aggregated payments summary for dashboard cards")
async def get_payments_summary(
    start_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):

    try:
        if start_date:
            start = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            if end_date:
                end = datetime.datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            else:

                end = datetime.datetime.now()
        else:

            today = datetime.datetime.now()
            start = today.replace(hour=0, minute=0, second=0, microsecond=0)
            end = today.replace(hour=23, minute=59, second=59, microsecond=999999)

        summary = await repository.get_payment_summary_by_methods(db, start, end)
        return {
            "start_date": start.strftime("%Y-%m-%d"),
            "end_date": end.strftime("%Y-%m-%d"),
            "summary": summary
        }
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный формат даты. Используйте YYYY-MM-DD")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения сводки: {str(e)}")

@router.post("/payments/create", summary="Создать тестовый платеж", tags=["Платежи"])
async def create_payment(
    payment: dict = Body(..., example={
        "type": "cash",
        "amount": 10000,
        "description": "Тестовый платеж",
        "timestamp": datetime.datetime.now().isoformat()
    }),
    db: AsyncSession = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):

    if isinstance(payment.get("timestamp"), str):
        payment["timestamp"] = datetime.datetime.fromisoformat(payment["timestamp"])
    new_payment = await repository.create_transaction(db, payment)
    return new_payment
